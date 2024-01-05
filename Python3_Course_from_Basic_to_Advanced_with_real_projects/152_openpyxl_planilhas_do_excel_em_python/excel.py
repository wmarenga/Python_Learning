"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""
import openpyxl
from random import uniform

"""
pedidos = openpyxl.load_workbook(
    '152_openpyxl_planilhas_do_excel_em_python/pedidos.xlsx') # workbook é um método (letra minúscula)

# Exibe os nomes da planilhas dentro de pedidos
nome_planilhas = pedidos.sheetnames
print(nome_planilhas)
planilha1 = pedidos['Página1']
print(planilha1)
print(planilha1['b4'].value)  # Exibe o valor na coluna 'b' e linha 4.

# Exibindo os valores da coluna 'b'
for campo in planilha1['b']:
    print(campo.value)

# Exibindo valores em um intervalo entre 'a1' e 'c2'
for linha in planilha1['a1:c2']:
    for coluna in linha:
        print(coluna.value)

# Exibe a planilha toda:
for linha in planilha1:
    for coluna in linha:
        print(coluna.value)

# Exibe toda a planilha coluna por coluna
for coluna in planilha1:
    for i in range(0, 1):
        if coluna[i].value is not None:
            print(coluna[i].value, end="\n")

# Alterando valores em uma célula da planilha
planilha1['b3'].value = 2200  # Altera mas ainda não salva
# Agora salva em uma nova planilha
pedidos.save('152_openpyxl_planilhas_do_excel_em_python/nova_planilha.xlsx')

# Inclui valores da linha cinco até a 15.
for linha in range(5, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido  # Coluna 1 ou "A"
    planilha1.cell(linha, 2).value = 1200 + linha  # Coluna 2 ou "B"

    # Preços aleatórios entre 10 e 100 com 2 casas decimais
    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco  # Coluna 3 ou "C"

pedidos.save('152_openpyxl_planilhas_do_excel_em_python/nova_planilha2.xlsx')

"""
# Agora queremos criar uma nova planilha

planilha = openpyxl.Workbook()  # Workbook() é uma classe (Letra Maiúscula)
# Criação das planilhas pelo índice (primeira posição)
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilha2.cell(
        linha, 1).value = f'Luiz {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(
        linha, 2).value = f'Otávio {linha} {round(uniform(10, 100), 2)}'
    planilha2.cell(
        linha, 3).value = f'Joãozinho {linha} {round(uniform(10, 100), 2)}'

planilha.save('152_openpyxl_planilhas_do_excel_em_python/nova_planilha.xlsx')
